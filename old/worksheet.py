from openpyxl import load_workbook
import re

import db

from paths import NG_FILEPATH, AG_FILEPATH

def check_sheet_exists(wb, sheet_name):
	try:
		wb[sheet_name]
		exists = True
	except KeyError as e:
		exists = False
		# print(e)
	finally:
		return exists
		
def create_sheets(wb, categories):		
	for sheet_name in categories:
		exists = check_sheet_exists(wb, sheet_name)
		if not exists:
			wb.create_sheet(title=sheet_name)
	wb.save(AG_FILEPATH)

def extract_data(filepath):
	wb = load_workbook(filepath)
	ws = wb['Main Goals']
	data_with_headers = ws["A1:B" + str(ws.max_row)]
	data = data_with_headers[1:]
	
	data = unique_data(data)
	
	return data, ws, wb
	
	
def extract_data_from_ws(ws):
	data_with_headers = ws["A1:B" + str(ws.max_row)]
	data = data_with_headers[1:]
	
	
def unique_data(data):
	return {(goal.value, cat.value) for (goal, cat) in data}
	

def add_new_goals(newpath=NG_FILEPATH, oldpath=AG_FILEPATH):
	data, _, _ = extract_data(newpath)
	data_old, ws, wb = extract_data(oldpath)
			
	items_to_add = [item for item in data if item not in data_old]
	
	if items_to_add:
		for item in items_to_add:	
			goal, cat = item
			ws.append({1: goal, 2: cat})
		wb.save(oldpath)				
		return True
	else:
		return False
		
def update_cat_goals(wb, cat, database):
	ws = wb[cat]	
	ws_goals = {ws.cell(row=1, column=col).value for col in range(1, ws.max_column+1)}
	
	db_goals = db.get_goals_from_cat(cat, database)
	for goal in db_goals:
		is_recorded = goal in ws_goals
		
		if is_recorded:
			continue
			
		else:
			no_goals_exist = list(ws_goals)[0]
			if no_goals_exist:
				ws.cell(row=1, column=ws.max_column).value = goal
				ws_goals = goal
				
			else:
				ws.cell(row=1, column=ws.max_column+1).value = goal
				
				ws_goals = {ws.cell(row=1, column=col).value for col in range(1, ws.max_column+1)}
		
def remove_goal_from_db(database, oldpath=AG_FILEPATH):
	data_trunc, ws, wb = extract_data(oldpath)
	
	cursor, conn = database
	data_db = db.fetch_all(cursor)
	
	for data in data_db:
		remains = data in list(data_trunc)
		
		if not remains:
			goal, _ = data
			t = (goal,)
			cursor.execute("DELETE FROM maingoals WHERE goal=?", t)
			
			
	conn.commit()
	data_db = db.fetch_all(cursor)
			
	
	
	
